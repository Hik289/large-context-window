"""
Excel document processor for .xls, .xlsx, and .csv files.
"""

from pathlib import Path
from typing import List
from agent_memory.core.segment import Segment
from agent_memory.processors.base_processor import FileProcessor, detect_file_type


class ExcelProcessor(FileProcessor):
    """Processor for Microsoft Excel files (.xls, .xlsx, .csv)."""

    def __init__(self, max_tokens_per_segment: int | None = None) -> None:
        self.max_tokens_per_segment = max_tokens_per_segment

    def can_process(self, file_path: Path) -> bool:
        """Return ``True`` for Excel/CSV files."""
        return detect_file_type(file_path) == "excel"

    def process(self, file_path: Path) -> List[Segment]:
        """Convert an Excel-like file into segments.

        Args:
            file_path: Path to the Excel/CSV file.

        Returns:
            Ordered list of :class:`Segment` objects extracted from the file.
        """
        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        if file_path.suffix.lower() == ".csv":
            return self._process_csv(file_path)
        return self._process_excel(file_path)

    def _process_csv(self, file_path: Path) -> List[Segment]:
        """Read a CSV file row-by-row into segments."""
        import csv

        segments: List[Segment] = []
        base_metadata = self._create_base_metadata(file_path, "excel")

        with open(file_path, "r", encoding="utf-8", newline="") as csvfile:
            # Sniff the dialect from a sample then rewind before reading.
            sample = csvfile.read(1024)
            csvfile.seek(0)
            dialect = csv.Sniffer().sniff(sample)

            reader = csv.reader(csvfile, dialect)
            rows = list(reader)

            if rows:
                # Treat the first row as the header.
                headers = rows[0]
                segments.append(
                    Segment(
                        content=", ".join(headers),
                        segment_type="table_header",
                        metadata={
                            **base_metadata,
                            "row_number": 1,
                            "column_count": len(headers),
                        },
                    )
                )

                # Remaining rows become individual data segments.
                for row_idx, row in enumerate(rows[1:], 2):
                    if not any(cell.strip() for cell in row):
                        continue
                    segments.append(
                        Segment(
                            content=", ".join(row),
                            segment_type="table_row",
                            metadata={
                                **base_metadata,
                                "row_number": row_idx,
                                "column_count": len(row),
                            },
                        )
                    )

        return segments

    def _process_excel(self, file_path: Path) -> List[Segment]:
        """Read an .xls/.xlsx workbook into per-worksheet segments."""
        try:
            import openpyxl
        except ImportError:
            raise ImportError(
                "openpyxl is required to process Excel files. Install with: pip install openpyxl"
            )

        segments: List[Segment] = []
        base_metadata = self._create_base_metadata(file_path, "excel")

        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                row_lines: List[str] = []
                non_empty_row_count = 0
                for row in sheet.iter_rows(values_only=True):
                    if not any(cell is not None and str(cell).strip() for cell in row):
                        continue
                    non_empty_row_count += 1
                    row_lines.append(
                        ", ".join(str(cell) if cell is not None else "" for cell in row)
                    )

                worksheet_metadata = {
                    **base_metadata,
                    "sheet_name": sheet_name,
                    "max_row": sheet.max_row,
                    "max_column": sheet.max_column,
                    "non_empty_row_count": non_empty_row_count,
                }

                if self.max_tokens_per_segment and self.max_tokens_per_segment > 0:
                    chunked_row_lines = self._chunk_rows_by_token_limit(
                        row_lines,
                        self.max_tokens_per_segment,
                    )

                    chunk_total = len(chunked_row_lines)
                    for chunk_index, chunk_lines in enumerate(chunked_row_lines, start=1):
                        segments.append(
                            Segment(
                                content="\n".join(chunk_lines),
                                segment_type="worksheet",
                                metadata={
                                    **worksheet_metadata,
                                    "worksheet_chunk_index": chunk_index,
                                    "worksheet_chunk_count": chunk_total,
                                },
                            )
                        )
                else:
                    # Whole sheet emitted as a single segment.
                    segments.append(
                        Segment(
                            content="\n".join(row_lines),
                            segment_type="worksheet",
                            metadata=worksheet_metadata,
                        )
                    )

            return segments

        except Exception as exc:
            raise ValueError(f"Failed to process Excel file: {exc}")

    def _chunk_rows_by_token_limit(
        self,
        row_lines: List[str],
        max_tokens: int,
    ) -> List[List[str]]:
        """Pack ``row_lines`` into chunks bounded by a whitespace-token cap."""
        chunks: List[List[str]] = []
        current_chunk: List[str] = []
        current_token_count = 0

        for row_line in row_lines:
            row_token_count = len(row_line.split())

            if not current_chunk:
                current_chunk.append(row_line)
                current_token_count = row_token_count
                continue

            # Open a fresh chunk if the row would push us over the limit.
            if current_token_count + row_token_count > max_tokens:
                chunks.append(current_chunk)
                current_chunk = [row_line]
                current_token_count = row_token_count
            else:
                current_chunk.append(row_line)
                current_token_count += row_token_count

        if current_chunk:
            chunks.append(current_chunk)

        return chunks
